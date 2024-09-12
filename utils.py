import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

def print_wb(xlsx, max_row=5):
    # Load the Excel file
    wb = load_workbook(xlsx, data_only=True)

    # Select the active worksheet
    ws = wb.active

    # Iterate over the rows and print the raw cell values
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):  # Adjust max_row as needed
        print(row)

        
def set_type_wbs(df):
    # force wbs codes to be string
    df['Cost center code'] =  df['Cost center code'].astype(int)
    df['Cost center code'] =  df['Cost center code'].astype(str)
    
    
def set_type_timestamp(df):
       # read these columns as datetime
    df['Booking start'] =  pd.to_datetime(df['Booking start'], format='%Y-%m-%d %H:%M')
    df['Booking end'] =  pd.to_datetime(df['Booking end'], format='%Y-%m-%d %H:%M')

    
re_timestamp = re.compile('[0-9]{8}-[0-9]{6}')

def save_invoice_with_timestamp(df, path):
    # remove old timestamp
    stem = re.sub(r'__[0-9]{8}-[0-9]{6}','', path.stem)
    
    # create new timestamp
    timestamp = str(datetime.now().strftime("%Y%m%d-%H%M%S"))
    
    output = path.parent / (stem + '__' + timestamp + '.xlsx')
    print('save_invoice_with_timestamp: ' + str(output))   
    df.to_excel(output, index=False)
    
    return output
    

def find_latest_invoice_version(invoice_path):
    re_version = re.compile(invoice_path.stem + '__[0-9]{8}-[0-9]{6}' + invoice_path.suffix)
    
    versions = []
    tmp = list(invoice_path.parent.glob("*" + invoice_path.suffix))
    pattern = invoice_path.stem + '__[0-9]{8}-[0-9]{6}' + invoice_path.suffix
    #print(pattern)
    for t in tmp:
        if re.search(pattern, str(t)):
            #print(t)
            versions.append(t)
    
    latest = invoice_path
    if len(versions) > 0:
        latest = sorted(versions)[-1]
    
    print('find_latest_invoice_version: ' + str(latest))   
    return latest


def check_totals(dataframe,tag,INVOICE_DIR,basename,ext='.xlsx'):
    _df = dataframe
    #print(_df.sum(numeric_only=True, axis=0))
    #print(_df.head()['Price'])
    
    _df['Charge'] = pd.to_numeric(_df['Charge'], errors='raise')
    #_df['Charge'] = _df['Charge'].round(2)
    
    # ungrouped total
    print("ungrouped: " + str(_df.sum(numeric_only=True, axis=0)['Charge'].round(2)))

    # totals by WBS
    tmp = _df.groupby(['Group','Remit code','Cost center code'])['Charge'].sum().reset_index()
    tmp.loc['Column_Total']= tmp.sum(numeric_only=True, axis=0)
    tmp.to_excel(INVOICE_DIR / ("test_" + basename + "__totals_by_group_and_wbs_" + tag + ext), index=False)
    total_wbs = round(tmp.loc['Column_Total']['Charge'],2)
    print("grouped by WBS: " + str(total_wbs))
    df_wbs = tmp.copy()

    # totals by instrument
    tmp = _df.groupby(['Resource/Product'])['Charge'].sum().reset_index()
    tmp.loc['Column_Total']= tmp.sum(numeric_only=True, axis=0)
    tmp.to_excel(INVOICE_DIR / ("test_ " + basename + "__totals_by_resource_" + tag  + ext), index=False)
    total_resource = round(tmp.loc['Column_Total']['Charge'],2)
    print("grouped by resource: " + str(total_resource))

    if total_resource != total_wbs:
        print("Totals don't match.")
        #return "Total don't match"
        
    return total_wbs, df_wbs
